import json
import sys
import os
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from extractor import extraer_datos_factura

def procesar_carpeta(ruta_carpeta: str):
    carpeta = Path(ruta_carpeta)
    
    if not carpeta.exists():
        print(f"❌ La carpeta '{ruta_carpeta}' no existe.")
        sys.exit(1)

    # Busca todos los PDF, JPG y PNG en la carpeta
    extensiones = ["*.pdf", "*.jpg", "*.jpeg", "*.png"]
    archivos = []
    for ext in extensiones:
        for archivo in carpeta.glob(ext):
            if archivo not in archivos:
                archivos.append(archivo)

    if not archivos:
        print(f"❌ No se encontraron facturas en '{ruta_carpeta}'.")
        sys.exit(1)

    print(f"\n📂 Carpeta: {ruta_carpeta}")
    print(f"📄 Facturas encontradas: {len(archivos)}\n")
    print("-" * 50)

    resultados = []

    for archivo in sorted(archivos):
        print(f"⏳ Procesando: {archivo.name} ...")
        try:
            datos = extraer_datos_factura(str(archivo))
            datos["archivo"] = archivo.name
            datos["estado"] = "OK"
            resultados.append(datos)
            print(f"   ✅ OK")
        except Exception as e:
            print(f"   ❌ Error: {e}")
            resultados.append({
                "archivo": archivo.name,
                "consumo_promedio_kwh": None,
                "estrato_socioeconomico": None,
                "referencia_transformador": None,
                "direccion": None,
                "numero_contrato": None,
                "estado": f"ERROR: {str(e)[:60]}"
            })

    print("-" * 50)
    print(f"\n✅ Procesadas: {sum(1 for r in resultados if r['estado'] == 'OK')}/{len(archivos)}")

    # Genera el nombre del archivo Excel con la fecha de hoy
    fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nombre_excel = f"resultados_{fecha}.xlsx"

    exportar_excel(resultados, nombre_excel)
    print(f"📊 Resultados guardados en: {nombre_excel}\n")

def exportar_excel(resultados: list, nombre_archivo: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas EPM"

    # Estilo encabezado
    color_verde = "1F7A4D"
    fuente_blanca = Font(color="FFFFFF", bold=True, size=11)
    relleno_verde = PatternFill(fill_type="solid", fgColor=color_verde)
    centrado = Alignment(horizontal="center", vertical="center")

    # Encabezados
    columnas = [
        ("Archivo",                 25),
        ("Consumo promedio (kWh)",  22),
        ("Estrato",                 10),
        ("Ref. Transformador",      20),
        ("Dirección",               40),
        ("Número de contrato",      22),
    ]

    for col_idx, (nombre, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre)
        celda.font = fuente_blanca
        celda.fill = relleno_verde
        celda.alignment = centrado
        ws.column_dimensions[celda.column_letter].width = ancho

    ws.row_dimensions[1].height = 25

    # Filas de datos
    color_fila_par = "F0FAF5"
    for fila_idx, dato in enumerate(resultados, start=2):
        valores = [
            dato.get("archivo"),
            dato.get("consumo_promedio_kwh"),
            dato.get("estrato_socioeconomico"),
            dato.get("referencia_transformador"),
            dato.get("direccion"),
            dato.get("numero_contrato"),
        ]
        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.alignment = Alignment(vertical="center", wrap_text=True)
            # Filas alternas con color suave
            if fila_idx % 2 == 0:
                celda.fill = PatternFill(fill_type="solid", fgColor=color_fila_par)
            # Si hubo error pinta la fila en rojo suave

        ws.row_dimensions[fila_idx].height = 20

    # Fila de totales al final
    fila_total = len(resultados) + 2
    ws.cell(row=fila_total, column=1, value="TOTAL FACTURAS").font = Font(bold=True)
    ws.cell(row=fila_total, column=2, value=len(resultados)).font = Font(bold=True)

    wb.save(nombre_archivo)

# ── Punto de entrada ──────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python procesar_lote.py <carpeta_con_facturas>")
        print("Ejemplo: python procesar_lote.py facturas/")
        sys.exit(1)

    procesar_carpeta(sys.argv[1])